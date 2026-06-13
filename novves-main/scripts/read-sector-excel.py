# -*- coding: utf-8 -*-
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

root = Path(__file__).resolve().parents[2]
path = root / "1.4 Sekt\u00f6rel Bazl\u0131 \u00dcr\u00fcn Gruplar\u0131 2026 R0.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out = {"sheets": {}}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    out["sheets"][name] = rows
wb.close()
out_path = Path(__file__).resolve().parent / "sector-excel-output.json"
out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print("written", out_path, "rows", sum(len(v) for v in out["sheets"].values()))
