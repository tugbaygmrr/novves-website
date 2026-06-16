#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import glob
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent
matches = list(ROOT.glob("1.1 Madde Kodlama*.xlsx"))
if not matches:
    print("Excel not found", file=sys.stderr)
    sys.exit(1)
xlsx = matches[0]
print("File:", xlsx.name)
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== {sn} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        print(i, row)
wb.close()
