#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
xlsx = list(ROOT.glob("1.1 Madde Kodlama*.xlsx"))[0]
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

sub_sheet = next(sn for sn in wb.sheetnames if "Alt Ailesi" in sn)
ws = wb[sub_sheet]
rows = list(ws.iter_rows(values_only=True))
wb.close()

header = None
models = []
for row in rows:
    if row[1] == "DÜZEY" and row[11] == "4.DÜZEY MADDE TANIMI":
        header = row
        continue
    if not header:
        continue
    level = row[1]
    try:
        level_n = int(level)
    except (TypeError, ValueError):
        continue
    if level_n not in (4, 5):
        continue
    family_code = str(row[8] or "").strip().upper()
    name_tr = str(row[11] or "").strip()
    if not family_code or family_code == "NA" or not name_tr or name_tr == "NA":
        continue
    models.append({
        "level": level_n,
        "category": str(row[6] or "").strip(),
        "family_code": family_code,
        "family_desc": str(row[9] or "").strip(),
        "sub_code": str(row[10] or "").strip(),
        "name_tr": name_tr,
        "product_code": str(row[12] or "").strip(),
        "short_code": str(row[13] or "").strip(),
        "short_code_tr": str(row[14] or "").strip(),
        "name_en": str(row[15] or "").strip() if row[15] else "",
    })

by_family = {}
for m in models:
    by_family.setdefault(m["family_code"], []).append(m)

(ROOT / "product-excel-models.json").write_text(
    json.dumps({"total": len(models), "by_family": by_family}, ensure_ascii=False, indent=2),
    encoding="utf-8",
)
print("total", len(models))
for k in sorted(by_family):
    print(k, len(by_family[k]))
