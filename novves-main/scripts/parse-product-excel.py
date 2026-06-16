#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
matches = list(ROOT.glob("1.1 Madde Kodlama*.xlsx"))
xlsx = matches[0]
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

out = {}

# Tematik sheet - families
ws = wb["Tematik"]
rows = list(ws.iter_rows(min_row=3, values_only=True))
families = []
for row in rows:
    if not row or not row[3]:
        continue
    families.append({
        "no": row[2],
        "en": str(row[3]).strip() if row[3] else "",
        "tr_name": str(row[4]).strip() if row[4] else "",
        "product_group_tr": str(row[5]).strip() if row[5] else "",
        "product_group_en": str(row[6]).strip() if row[6] else "",
        "inspiration": str(row[7]).strip() if row[7] else "",
        "description": str(row[8]).strip() if row[8] else "",
    })
out["families"] = families

# Ana Ürün Gruplar? ve Aileleri
for sn in wb.sheetnames:
    if "Ana" in sn and "r" in sn.lower():
        ws = wb[sn]
        rows = list(ws.iter_rows(max_row=20, values_only=True))
        out["ana_sheet"] = sn
        out["ana_preview"] = [list(r) for r in rows]
        all_rows = list(ws.iter_rows(values_only=True))
        out["ana_row_count"] = len(all_rows)
        out["ana_rows"] = [list(r) for r in all_rows[:50]]

# Ürün Alt Ailesi
for sn in wb.sheetnames:
    if "Alt Ailesi" in sn or "Alt Ailes" in sn:
        ws = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        out["sub_sheet"] = sn
        out["sub_row_count"] = len(all_rows)
        out["sub_preview"] = [list(r) for r in all_rows[:25]]
        models = []
        for row in all_rows[3:]:
            if not row:
                continue
            vals = [str(c).strip() if c is not None else "" for c in row]
            if any(vals):
                models.append(vals)
        out["sub_models_count"] = len(models)
        out["sub_models_sample"] = models[:30]

wb.close()

out_path = ROOT / "product-excel-parsed.json"
out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print("wrote", out_path)
print("families", len(families))
print("sub_models", out.get("sub_models_count", 0))
